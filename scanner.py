import tkinter as tk
import cv2
import openpyxl
import openpyxl.workbook
from PIL import Image, ImageTk

list_of_data = []
list_of_qr = []
list_of_name = []
list_of_read = []
last_row = 0

def resize_image(image_path, new_width, new_height):
    image = Image.open(image_path)
    resized_image = image.resize((new_width, new_height))
    return resized_image

def show_valid():
    window = tk.Tk()
    window.title("My GUI Window")
    window.geometry("400x400")
    image_path = "check.png"  # Replace with the actual path
    resized_image = resize_image(image_path, 300, 300)
    photo = ImageTk.PhotoImage(resized_image)
    label = tk.Label(window, image=photo)
    label.pack()

    label = tk.Label(window, text="Your QR code is valid!")
    label.pack()

    window.after(2500, window.destroy)
    window.mainloop()

def show_invalid():
    window = tk.Tk()
    window.title("My GUI Window")
    window.geometry("400x400")
    image_path = "cross.png"  # Replace with the actual path
    resized_image = resize_image(image_path, 300, 300)
    photo = ImageTk.PhotoImage(resized_image)
    label = tk.Label(window, image=photo)
    label.pack()

    label = tk.Label(window, text="This QR code is already Registered.")
    label.pack()

    window.after(3000, window.destroy)
    window.mainloop()

def create_excel_file():
    create_excel = openpyxl.Workbook()
    create_excel.save('datatest.xlsx')

def add_data_to_excel(name,code,stat):
    workbook = openpyxl.load_workbook('datatest.xlsx')

    sheet = workbook['Sheet']

    last_row = sheet.max_row + 1

    sheet.cell(row=last_row, column=1).value = name
    sheet.cell(row=last_row, column=2).value = code
    sheet.cell(row=last_row, column=3).value = stat

    workbook.save('datatest.xlsx')

def name_from_excel(data):
    workbook = openpyxl.load_workbook('datatest.xlsx')

    sheet = workbook['Sheet']

    last_row = sheet.max_row + 1
    info = read_data_from_excel("data.xlsx")
    for i in range(1,len(info)):
        if info[i][-1] == data:
            name = info[i][2]
            return name

    sheet.cell(row=last_row, column=1).value = name

    workbook.save('datatest.xlsx')

def read_data_from_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
    return data

window_width = int(800*1.15)
window_height = int(600*1.15)

def scan_multiple_qr_codes():
    cap = cv2.VideoCapture(0)
    detector = cv2.QRCodeDetector()

    while True:
        ret, frame = cap.read()

        if not ret:
            break

        data, _, _ = detector.detectAndDecode(frame)

        if data:
            list_of_qr.append(data)
            name = name_from_excel(data)
            if data in list_of_data and name not in list_of_read:
                print("VALID")
                show_valid()
                add_data_to_excel(name,data, "VALID")
                list_of_data.remove(data)
            else:
                print("INVALID")
                show_invalid()

        frame = cv2.resize(frame, (window_width, window_height))

        cv2.imshow("QR Code Scanner", frame)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    cap.release()
    cv2.destroyAllWindows()

infodata = read_data_from_excel("data.xlsx")
for i in range(1,len(infodata)):
    list_of_data.append(infodata[i][-1])

inforeader = read_data_from_excel("datatest.xlsx")
for i in range(1,len(inforeader)):
    list_of_read.append(inforeader[i][0])

if __name__ == "__main__":scan_multiple_qr_codes()

